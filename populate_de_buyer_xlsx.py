from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook


APPEND_HEADERS = [
    "Matched Item No.",
    "Matched Mfr Catalog No.",
    "Matched Brand Name",
    "Matched Item Description",
    "Matched Image Link",
    "Matched Overview",
    "Matched Length",
    "Matched Width",
    "Matched Height",
    "Matched Volume",
    "Matched Diameter",
    "Matched Color",
    "Matched Material",
    "Matched EAN Code",
    "Matched Pattern",
    "Matched Barcode",
    "Matched Product URL",
    "Matched Price",
    "Matched Source File",
]

JSON_FIELD_ORDER = [
    "item_no",
    "mfr_catalog_no",
    "brand_name",
    "item_description",
    "image_link",
    "overview",
    "length",
    "width",
    "height",
    "volume",
    "diameter",
    "color",
    "material",
    "ean_code",
    "pattern",
    "barcode",
    "product_url",
    "price",
    "source_file",
]

JSON_KEY_ALIASES = {
    "item_no": ["item_no", "item_no.", "Item No."],
    "mfr_catalog_no": ["mfr_catalog_no", "mfr_catalog_no.", "Mfr Catalog No."],
    "brand_name": ["brand_name", "brand", "Brand Name"],
    "item_description": ["item_description", "description", "name", "Item Description"],
    "image_link": ["image_link", "Image Link"],
    "overview": ["overview", "Overview"],
    "length": ["length", "Length"],
    "width": ["width", "Width"],
    "height": ["height", "Height"],
    "volume": ["volume", "Volume"],
    "diameter": ["diameter", "Diameter"],
    "color": ["color", "Color"],
    "material": ["material", "Material"],
    "ean_code": ["ean_code", "EAN Code"],
    "pattern": ["pattern", "Pattern"],
    "barcode": ["barcode", "Barcode"],
    "product_url": ["product_url", "url", "Product URL"],
    "price": ["price", "Price"],
}


def normalize_mfr(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").split()).upper()


def clean_cell_value(value: Any) -> Any:
    if isinstance(value, str):
        value = value.strip()
        return value if value != "" else None
    return value


def first_non_empty(record: Dict[str, Any], keys: List[str]) -> Any:
    for key in keys:
        if key in record:
            value = clean_cell_value(record.get(key))
            if value is not None:
                return value
    return None


def record_from_json_entry(entry: Dict[str, Any], source_file: str) -> Dict[str, Any]:
    record: Dict[str, Any] = {}
    for field, aliases in JSON_KEY_ALIASES.items():
        record[field] = first_non_empty(entry, aliases)
    if record.get("brand_name") is None and record.get("item_description"):
        record["brand_name"] = "de Buyer"
    record["source_file"] = source_file
    return record


def iter_json_records(json_dir: Path) -> Iterable[Dict[str, Any]]:
    for path in sorted(json_dir.glob("*.json")):
        if path.name == "all_products.json":
            continue
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        if not isinstance(data, list):
            continue
        for entry in data:
            if isinstance(entry, dict):
                yield record_from_json_entry(entry, path.name)


def iter_workbook_rows(sheet) -> Iterable[Dict[str, Any]]:
    for row in sheet.iter_rows(min_row=1, values_only=True):
        item_no = clean_cell_value(row[0] if len(row) > 0 else None)
        mfr_catalog_no = clean_cell_value(row[1] if len(row) > 1 else None)
        brand_name = clean_cell_value(row[2] if len(row) > 2 else None)
        if mfr_catalog_no is None:
            continue
        yield {
            "item_no": item_no,
            "mfr_catalog_no": mfr_catalog_no,
            "brand_name": brand_name,
            "source_file": "all_products.xlsx",
        }


def build_product_lookup(json_dir: Path, workbook_sheet=None) -> Dict[str, Dict[str, Any]]:
    lookup: Dict[str, Dict[str, Any]] = {}

    for record in iter_json_records(json_dir):
        key = normalize_mfr(record.get("mfr_catalog_no"))
        if not key or key in lookup:
            continue
        lookup[key] = record

    if workbook_sheet is not None:
        for record in iter_workbook_rows(workbook_sheet):
            key = normalize_mfr(record.get("mfr_catalog_no"))
            if not key or key in lookup:
                continue
            lookup[key] = record

    return lookup


def record_to_row(record: Dict[str, Any]) -> List[Any]:
    return [record.get(field) for field in JSON_FIELD_ORDER]


def populate_source_sheet(source_sheet, product_lookup: Dict[str, Dict[str, Any]]) -> int:
    start_col = 15  # Column O
    for offset, header in enumerate(APPEND_HEADERS):
        source_sheet.cell(row=1, column=start_col + offset, value=header)

    matched_rows = 0
    for row_index in range(2, source_sheet.max_row + 1):
        source_mfr = normalize_mfr(source_sheet.cell(row=row_index, column=2).value)
        if not source_mfr:
            continue

        product_row = product_lookup.get(source_mfr)
        if product_row is None:
            continue

        matched_rows += 1
        for offset, value in enumerate(record_to_row(product_row)):
            source_sheet.cell(row=row_index, column=start_col + offset, value=value)

    for offset, width in enumerate((16, 18, 24, 36, 28, 48, 14, 14, 14, 14, 14, 14, 18, 18, 14, 18, 32, 14, 22), start=start_col):
        source_sheet.column_dimensions[source_sheet.cell(row=1, column=offset).column_letter].width = width

    return matched_rows


def resolve_default_paths(base_dir: Path) -> Tuple[Path, Path, Path]:
    source_path = base_dir / "sources" / "DE BUYER.xlsx"
    product_path = base_dir / "debuyer" / "all_products.xlsx"
    output_path = base_dir / "sources" / "DE BUYER_populated.xlsx"
    return source_path, product_path, output_path


def main() -> int:
    base_dir = Path(__file__).resolve().parent
    default_source, default_products, default_output = resolve_default_paths(base_dir)

    parser = argparse.ArgumentParser(
        description="Populate DE BUYER.xlsx with matched product data from all_products.xlsx without modifying the original workbook."
    )
    parser.add_argument("--source", type=Path, default=default_source, help="Path to DE BUYER.xlsx")
    parser.add_argument("--products", type=Path, default=default_products, help="Path to all_products.xlsx")
    parser.add_argument("--output", type=Path, default=default_output, help="Path for the populated workbook")
    args = parser.parse_args()

    if not args.source.exists():
        raise FileNotFoundError(f"Source workbook not found: {args.source}")
    if not args.products.exists():
        raise FileNotFoundError(f"Products workbook not found: {args.products}")

    source_workbook = load_workbook(args.source)
    product_workbook = load_workbook(args.products, read_only=True, data_only=True)

    try:
        source_sheet = source_workbook.active
        product_sheet = product_workbook.active
        product_lookup = build_product_lookup(base_dir / "debuyer", product_sheet)
        matched_rows = populate_source_sheet(source_sheet, product_lookup)

        args.output.parent.mkdir(parents=True, exist_ok=True)
        source_workbook.save(args.output)
    finally:
        product_workbook.close()

    print(f"Matched {matched_rows} source rows and wrote {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())